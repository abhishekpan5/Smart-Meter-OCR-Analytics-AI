#!/usr/bin/env python3
"""
Migration script to fix the smart_meter_database.db schema and reload data from meter_ref_data.xlsx
"""

import sqlite3
import shutil
from pathlib import Path
from openpyxl import load_workbook

DB_PATH = Path("smart_meter_database.db")
EXCEL_PATH = Path("meter_ref_data.xlsx")
BACKUP_PATH = Path("smart_meter_database_backup.db")

# 1. Backup the existing database
if DB_PATH.exists():
    print(f"Backing up {DB_PATH} to {BACKUP_PATH}")
    shutil.copy(DB_PATH, BACKUP_PATH)
else:
    print(f"Database file {DB_PATH} does not exist. Exiting.")
    exit(1)

# 2. Connect to the database
conn = sqlite3.connect(str(DB_PATH))
cursor = conn.cursor()

# 3. Drop the old table if it exists
print("Dropping old meter_readings table if it exists...")
cursor.execute("DROP TABLE IF EXISTS meter_readings")

# 4. Create the new table with the correct schema
print("Creating new meter_readings table...")
cursor.execute('''
    CREATE TABLE meter_readings (
        meter_serial_number TEXT PRIMARY KEY,
        meter_reading TEXT NOT NULL,
        reading_unit TEXT,
        reading_date TEXT
    )
''')

# 5. Load data from Excel
print(f"Loading data from {EXCEL_PATH}...")
wb = load_workbook(EXCEL_PATH)
ws = wb.active

# Get headers and their column indices
headers = [cell.value for cell in ws[1]]
header_indices = {header: idx for idx, header in enumerate(headers)}

required_headers = ["meter_serial_number", "meter_reading", "reading unit", "reading_date"]
for h in required_headers:
    if h not in header_indices:
        print(f"Missing required column in Excel: {h}")
        conn.close()
        exit(1)

row_count = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    if not any(row):
        continue
    meter_serial_number = row[header_indices["meter_serial_number"]]
    meter_reading = row[header_indices["meter_reading"]]
    reading_unit = row[header_indices["reading unit"]]
    reading_date = row[header_indices["reading_date"]]

    # Ensure meter_reading is a string and preserve leading zeros
    meter_reading_str = str(meter_reading).zfill(len(str(meter_reading))) if isinstance(meter_reading, int) else str(meter_reading)

    # Convert reading_date to string if it's a datetime
    if hasattr(reading_date, 'isoformat'):
        reading_date_str = reading_date.isoformat(sep=' ')
    else:
        reading_date_str = str(reading_date)

    # Debug: print row values and types
    print(f"Inserting row: meter_serial_number={meter_serial_number} ({type(meter_serial_number)}), meter_reading='{meter_reading_str}' ({type(meter_reading_str)}), reading_unit='{reading_unit}' ({type(reading_unit)}), reading_date='{reading_date_str}' ({type(reading_date_str)})")

    cursor.execute('''
        INSERT INTO meter_readings (meter_serial_number, meter_reading, reading_unit, reading_date)
        VALUES (?, ?, ?, ?)
    ''', (meter_serial_number, meter_reading_str, reading_unit, reading_date_str))
    row_count += 1

conn.commit()
conn.close()
print(f"Migration complete. {row_count} rows loaded into meter_readings table.") 